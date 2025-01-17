from copy import deepcopy
from types import MethodType
import os
import openpyxl
import sys
import time
import traceback
import datetime
import statistics
class Ref:
    def __init__(self,ws):
        # Initialisations
        self.score=0
        self.longueur_chute=0
        self.nombre_chute=0
        self.longueur_chute_commande=0
        self.longueur_chute_chute=0
        self.longueur_chute_chute=0
        self.result=[]
        self.reference=""
        self.commandes=[]
        self.chutes=[]
        self.stocks=[]
        self.stocks_finaux=[]
        self.ws=None
        # Récupération des arguments
        self.ws=ws
    def generate_output(self):
        # Génere le résultat
        for i in range(len(self.result)):
            self.ws.cell(i+4,11,i)
            self.ws.cell(i+4,12,self.result[i][1])
            self.ws.cell(i+4,13,self.result[i][0])
            self.ws.cell(i+4,14,self.result[i][2])
        for i in range(len(self.result),len(self.result)+100):
            self.ws.cell(4+i,11,"")
            self.ws.cell(4+i,12,"")
            self.ws.cell(4+i,13,"")
            self.ws.cell(4+i,14,"")
        # Génere le résumé
        self.ws.cell(4,8,"nb commandes")
        self.ws.cell(4,9,len(self.commandes))
        self.ws.cell(4,10,"commandes")
        self.ws.cell(5,8,"nb stocks")
        self.ws.cell(5,9,len(self.stocks))
        self.ws.cell(5,10,"stocks")
        self.ws.cell(6,8,"nb chutes")
        self.ws.cell(6,9,len(self.chutes))
        self.ws.cell(6,10,"chutes")
        self.ws.cell(7,8,"longueur commandes")
        self.ws.cell(7,9,sum(self.commandes))
        self.ws.cell(7,10,"mm")
        self.ws.cell(8,8,"longueur stocks")
        self.ws.cell(8,9,sum(self.stocks))
        self.ws.cell(8,10,"mm")
        self.ws.cell(9,8,"longueur chutes")
        self.ws.cell(9,9,sum(self.chutes))
        self.ws.cell(9,10,"mm")
        self.ws.cell(10,8,"chutes par commandes")
        self.ws.cell(10,9,int(sum(self.chutes)/len(self.commandes)))
        self.ws.cell(10,10,"mm / commandes")
        self.ws.cell(11,8,"longueur par chutes")
        self.ws.cell(11,9,int(sum(self.chutes)/(len(self.chutes))))
        self.ws.cell(11,10,"mm / chute")
        # Génère le resultat des chutes
        self.ws.cell(12,8,"numero chute")
        self.ws.cell(12,9,"longueur chute (mm)")
        self.ws.cell(12,10,"% chute")
        for i in range(len(self.chutes)):
            self.ws.cell(13+i,8,i)
            self.ws.cell(13+i,9,self.chutes[i])
            self.ws.cell(13+i,10,int(100*self.chutes[i]/sum(self.chutes)))
        for i in range(len(self.chutes),len(self.chutes)+100):
            self.ws.cell(13+i,8,"")
            self.ws.cell(13+i,9,"")
            self.ws.cell(13+i,10,"")
    def get_commandes(self):
        i=3
        while 1:
            i+=1
            try:
                if self.ws.cell(i,1).value==None:
                    break
                else:
                    for j in range(self.ws.cell(i,2).value):
                        self.commandes.append(self.ws.cell(i,1).value)
            except:
                break
    def get_stocks(self):
        i=3
        while 1:
            i+=1
            try:
                if self.ws.cell(i,3).value==None:
                    break
                else:
                    for j in range(self.ws.cell(i,4).value):
                        self.stocks.append(self.ws.cell(i,3).value)
            except:
                break
    def get_setting(self,setting):
        i=3
        while 1:
            i+=1
            try:
                if self.ws.cell(i,5).value==None:
                    break
                elif self.ws.cell(i,5).value==setting:
                    exec("self."+setting+"="+str(self.ws.cell(i,6).value))
            except:
                break
    def process(self):
        min_chutes=sum(self.commandes)
        i_min=0
        i=8
        temp_commandes=deepcopy(self.commandes)
        temp_stocks=deepcopy(self.stocks)
        temp_stocks_finaux=deepcopy(self.stocks)
        temp_result=deepcopy(self.result)
        temp_chutes=deepcopy(self.chutes)
        temp_results0=[]
        temp_results1=[]
        temp_results2=[]
        best_stocks_finaux=deepcopy(self.stocks)
        best_result=deepcopy(self.result)
        best_chutes=deepcopy(self.chutes)
        while 1:
            try:
                print("_"*100)
                self.commandes=deepcopy(temp_commandes)
                self.stocks=deepcopy(temp_stocks)
                self.stocks_finaux=deepcopy(temp_stocks)
                self.result=deepcopy(temp_result)
                self.chutes=deepcopy(temp_chutes)
                while len(self.commandes)>0:
                    self.stocks_finaux.sort(reverse=False)
                    command="self.f"+str(i)+"()"
                    exec(command)
                temp_results0.append("f"+str(i)+": "+str(sum(self.chutes)))
                temp_results1.append("f"+str(i)+": "+str(int(sum(self.chutes)/(len(self.chutes)))))
                temp_results2.append("f"+str(i)+": "+str(self.chutes))
                if sum(self.chutes)>0:
                    if min_chutes>sum(self.chutes):
                        min_chutes=sum(self.chutes)
                        i_min=i
                        best_stocks_finaux=deepcopy(self.stocks)
                        best_result=deepcopy(self.result)
                        best_chutes=deepcopy(self.chutes)
            except AttributeError:
                break
            except:
                traceback.print_exc()
                time.sleep(5)
            i=i+1
        print("_"*100)
        self.commandes=deepcopy(temp_commandes)
        self.stocks=deepcopy(temp_stocks)
        self.stocks_finaux=deepcopy(best_stocks_finaux)
        self.result=deepcopy(best_result)
        self.chutes=deepcopy(best_chutes)
        
        print(temp_results0)
        print(temp_results1)
        for i in range(len(temp_results2)):
            print(str(temp_results2[i]))
    
    
    def cut(self,stock,commande):
        self.cut_index(self.stocks.index(stock),self.commandes.index(commande))
    
    def cut_index(self,index_stock,index_commande):
        
        print()
        print("commandes: "+str(self.commandes))
        print("stocks_debut: "+str(self.stocks_finaux))
        
        commande=self.commandes[index_commande]
        stock=self.stocks_finaux[index_stock]
        final=stock-commande-self.longueur_decoupe
        if final<min(self.commandes):
            self.chutes.append(final)
        self.result.append([stock,commande,final])
        self.stocks_finaux[index_stock]=final
        self.commandes.remove(commande)
        
        print("stocks_fin: "+str(self.stocks_finaux))
        print("chutes: "+str(self.chutes))
        print()
        print("index_commande: "+str(index_commande))
        print("index_stock: "+str(index_stock))
        print("stock: "+str(stock))
        print("commande: "+str(commande))
        print("final: "+str(final))
        print("-"*50)
        if final<0:
            time.sleep(2)
            exit()
    def f0(self):  # Minimum commande
        print(sys._getframe().f_code.co_name)
        index_stock=0
        index_commande=self.commandes.index(min(self.commandes))
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        self.cut_index(index_stock,index_commande)
    def f1(self):  # Maximum commande
        print(sys._getframe().f_code.co_name)
        index_stock=0
        index_commande=self.commandes.index(max(self.commandes))
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        self.cut_index(index_stock,index_commande)
    def f2(self):  # Medianne commande
        print(sys._getframe().f_code.co_name)
        index_stock=0
        index_commande=self.commandes.index(self.commandes[len(self.commandes)//2])
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        self.cut_index(index_stock,index_commande)
    def f3(self):  # Moyenne commande
        print(sys._getframe().f_code.co_name)
        index_stock=0
        moyenne=sum(self.commandes)/len(self.commandes)
        for i in range(len(self.commandes)):
            commande=self.commandes[i]
            if commande>=moyenne:
                index_commande=i
                break
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        self.cut_index(index_stock,index_commande)
    def f4(self):  # Mini final avec 1 coup d'avance (si chute)
        print(sys._getframe().f_code.co_name)
        index_stock=0
        commande=min(self.commandes)
        index_commande=self.commandes.index(commande)
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        
        if len(self.commandes)>2:
            stock=self.stocks_finaux[index_stock]
            final1=stock-commande-self.longueur_decoupe
            
            commandes=deepcopy(self.commandes)
            commandes.remove(commande)
            commande2=min(commandes)
            
            final2=final1-commande2-self.longueur_decoupe
            
            
            print(str([final2,final1,stock]))
            print(str([commande2,commande]))
            
            final_min=final2
            finaux=[commande,commande2]
            if final2<min(self.commandes) and final2>0:
                for i in range(len(self.commandes)-1):
                    for j in range(i+1,len(self.commandes)):
                        final=stock-self.commandes[i]-self.commandes[j]-2*self.longueur_decoupe
                        
                        print("final: "+str(final))
                        print("final_min: "+str(final_min))
                    
                        if final>0 and final<final_min:
                            final_min=final
                            finaux=deepcopy([self.commandes[i],self.commandes[j]])
                         
                commande1=finaux[0]
                commande2=finaux[1]
                
                print()
                print("Double")
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande1)
                print()
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande2)
                print("fin double")
            else:
                self.cut_index(index_stock,index_commande)
        else:
            self.cut_index(index_stock,index_commande)
    def f5(self):  # Maxi final avec 1 coup d'avance (si chute)
        print(sys._getframe().f_code.co_name)
        index_stock=0
        commande=max(self.commandes)
        index_commande=self.commandes.index(commande)
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        
        if len(self.commandes)>2:
            stock=self.stocks_finaux[index_stock]
            final1=stock-commande-self.longueur_decoupe
            
            commandes=deepcopy(self.commandes)
            commandes.remove(commande)
            commande2=min(commandes)
            
            final2=final1-commande2-self.longueur_decoupe
            
            
            print(str([final2,final1,stock]))
            print(str([commande2,commande]))
            
            final_min=final2
            finaux=[commande,commande2]
            if final2<min(self.commandes) and final2>0:
                for i in range(len(self.commandes)-1):
                    for j in range(i+1,len(self.commandes)):
                        final=stock-self.commandes[i]-self.commandes[j]-2*self.longueur_decoupe
                        
                        print("final: "+str(final))
                        print("final_min: "+str(final_min))
                    
                        if final>0 and final<final_min:
                            final_min=final
                            finaux=deepcopy([self.commandes[i],self.commandes[j]])
                         
                commande1=finaux[0]
                commande2=finaux[1]
                
                print()
                print("Double")
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande1)
                print()
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande2)
                print("fin double")
            else:
                self.cut_index(index_stock,index_commande)
        else:
            self.cut_index(index_stock,index_commande)
    def f6(self):  # Medianne final avec 1 coup d'avance (si chute)
        print(sys._getframe().f_code.co_name)
        index_stock=0
        commande=self.commandes[len(self.commandes)//2]
        index_commande=self.commandes.index(commande)
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        
        if len(self.commandes)>2:
            stock=self.stocks_finaux[index_stock]
            final1=stock-commande-self.longueur_decoupe
            
            commandes=deepcopy(self.commandes)
            commandes.remove(commande)
            commande2=min(commandes)
            
            final2=final1-commande2-self.longueur_decoupe
            
            
            print(str([final2,final1,stock]))
            print(str([commande2,commande]))
            
            final_min=final2
            finaux=[commande,commande2]
            if final2<min(self.commandes) and final2>0:
                for i in range(len(self.commandes)-1):
                    for j in range(i+1,len(self.commandes)):
                        final=stock-self.commandes[i]-self.commandes[j]-2*self.longueur_decoupe
                        
                        print("final: "+str(final))
                        print("final_min: "+str(final_min))
                    
                        if final>0 and final<final_min:
                            final_min=final
                            finaux=deepcopy([self.commandes[i],self.commandes[j]])
                         
                commande1=finaux[0]
                commande2=finaux[1]
                
                print()
                print("Double")
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande1)
                print()
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande2)
                print("fin double")
            else:
                self.cut_index(index_stock,index_commande)
        else:
            self.cut_index(index_stock,index_commande)
    def f7(self):  # Moyenne final avec 1 coup d'avance (si chute)
        print(sys._getframe().f_code.co_name)
        index_stock=0
        
        moyenne=sum(self.commandes)/len(self.commandes)
        for i in range(len(self.commandes)):
            commande=self.commandes[i]
            if commande>=moyenne:
                index_commande=i
                break
        commande=self.commandes[index_commande]
        index_commande=self.commandes.index(commande)
        while self.stocks_finaux[index_stock]<self.commandes[index_commande]:
            index_stock+=1
        
        if len(self.commandes)>2:
            stock=self.stocks_finaux[index_stock]
            final1=stock-commande-self.longueur_decoupe
            
            commandes=deepcopy(self.commandes)
            commandes.remove(commande)
            commande2=min(commandes)
            
            final2=final1-commande2-self.longueur_decoupe
            
            
            print(str([final2,final1,stock]))
            print(str([commande2,commande]))
            
            final_min=final2
            finaux=[commande,commande2]
            if final2<min(self.commandes) and final2>0:
                for i in range(len(self.commandes)-1):
                    for j in range(i+1,len(self.commandes)):
                        final=stock-self.commandes[i]-self.commandes[j]-2*self.longueur_decoupe
                        
                        print("final: "+str(final))
                        print("final_min: "+str(final_min))
                    
                        if final>0 and final<final_min:
                            final_min=final
                            finaux=deepcopy([self.commandes[i],self.commandes[j]])
                         
                commande1=finaux[0]
                commande2=finaux[1]
                
                print()
                print("Double")
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande1)
                print()
                print("stock: "+str(stock))
                print("commande1: "+str(commande1))
                print("commande2: "+str(commande2))
                print()
                self.cut(stock,commande2)
                print("fin double")
            else:
                self.cut_index(index_stock,index_commande)
        else:
            self.cut_index(index_stock,index_commande)
    
    def f8(self):
        print(sys._getframe().f_code.co_name)
        
    
chemin=str(os.path.dirname(os.path.abspath(__file__)))+"\\"
document="Data.xlsx"
wb=openpyxl.load_workbook(chemin+document)
references=wb.sheetnames
for reference in references:
    ref=Ref(wb[reference])
    ref.get_commandes()
    ref.get_stocks()
    ref.get_setting("limite_chute")
    ref.get_setting("longueur_decoupe")
    ref.process()
    ref.generate_output()
time_expiration=1565691296
time_left=time_expiration-time.time()
print("_"*100)
print("Version d'essai")
print("Vincent Bénet")
print("Ingénieur Arts & Métiers")
print("Expiration: "+str(datetime.datetime.fromtimestamp(time_expiration).strftime('%Y-%m-%d %H:%M:%S')))
if time_left<0:
    print("version périmée")
    print("Pour continuer appeller:")
    print("+33640123744")
else:
    print("_"*100)
    print("Document disponible à: "+chemin+document)
    wb.save(filename=chemin+document)